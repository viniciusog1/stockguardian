"""Unit: helpers ExportFile (nome/media type/bytes válidos)."""

from __future__ import annotations

import io
import uuid
from datetime import UTC, datetime
from decimal import Decimal

import pytest
from app.schemas.report import (
    InventoryValuationItem,
    InventoryValuationReport,
    InventoryValuationSummary,
)
from app.utils.excel import XLSX_MEDIA_TYPE, inventory_valuation_export_file
from openpyxl import load_workbook

pytestmark = pytest.mark.unit


def _report() -> InventoryValuationReport:
    return InventoryValuationReport(
        generated_at=datetime(2026, 6, 26, tzinfo=UTC),
        summary=InventoryValuationSummary(
            total_products=1, total_units=5, total_value=Decimal("50.00")
        ),
        items=[
            InventoryValuationItem(
                product_id=uuid.uuid4(),
                sku="A-1",
                name="Produto A",
                quantity=5,
                unit_price=Decimal("10.00"),
                stock_value=Decimal("50.00"),
            )
        ],
    )


def test_export_file_fields_and_valid_xlsx() -> None:
    export = inventory_valuation_export_file(_report())
    assert export.filename == "inventory-valuation-2026-06-26.xlsx"
    assert export.media_type == XLSX_MEDIA_TYPE
    wb = load_workbook(io.BytesIO(export.content))
    assert "Valuation" in wb.sheetnames
