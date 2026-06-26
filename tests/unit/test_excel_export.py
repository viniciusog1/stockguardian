"""Testes unitários da serialização Excel dos relatórios (sem DB)."""

from __future__ import annotations

import io
import uuid
from datetime import UTC, datetime
from decimal import Decimal

import pytest
from app.models.stock_movement import MovementType
from app.schemas.report import (
    InventoryValuationItem,
    InventoryValuationReport,
    InventoryValuationSummary,
    MovementsSummaryReport,
    MovementsSummaryRow,
)
from app.utils.excel import (
    inventory_valuation_workbook,
    movements_summary_workbook,
    workbook_to_bytes,
)
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

pytestmark = pytest.mark.unit


def _valuation_report() -> InventoryValuationReport:
    items = [
        InventoryValuationItem(
            product_id=uuid.uuid4(),
            sku="A-1",
            name="Produto A",
            quantity=5,
            unit_price=Decimal("10.00"),
            stock_value=Decimal("50.00"),
        ),
    ]
    return InventoryValuationReport(
        generated_at=datetime(2026, 6, 26, tzinfo=UTC),
        summary=InventoryValuationSummary(
            total_products=1, total_units=5, total_value=Decimal("50.00")
        ),
        items=items,
    )


def _movements_report() -> MovementsSummaryReport:
    rows = [
        MovementsSummaryRow(type=MovementType.IN, movement_count=2, total_quantity=15),
        MovementsSummaryRow(type=MovementType.OUT, movement_count=1, total_quantity=3),
        MovementsSummaryRow(type=MovementType.ADJUSTMENT, movement_count=0, total_quantity=0),
    ]
    return MovementsSummaryReport(
        generated_at=datetime(2026, 6, 26, tzinfo=UTC),
        date_from=None,
        date_to=None,
        rows=rows,
        total_movements=3,
    )


def _cells(ws: Worksheet) -> list[object]:
    return [c.value for row in ws.iter_rows() for c in row]


def test_valuation_workbook_has_item_and_totals() -> None:
    wb = inventory_valuation_workbook(_valuation_report())
    ws = wb.active
    assert isinstance(ws, Worksheet)
    values = _cells(ws)
    assert "SKU" in values
    assert "A-1" in values
    assert "Produto A" in values
    # money fica como Decimal no workbook em memória (vira float ao salvar/reabrir)
    numeric = [float(v) for v in values if isinstance(v, int | float | Decimal)]
    assert 50.0 in numeric  # stock_value do item / total_value


def test_movements_workbook_has_three_type_rows() -> None:
    wb = movements_summary_workbook(_movements_report())
    ws = wb.active
    assert isinstance(ws, Worksheet)
    values = [str(v) for v in _cells(ws) if v is not None]
    assert "in" in values and "out" in values and "adjustment" in values


def test_workbook_to_bytes_roundtrips() -> None:
    wb = inventory_valuation_workbook(_valuation_report())
    data = workbook_to_bytes(wb)
    assert isinstance(data, bytes) and len(data) > 0
    reopened = load_workbook(io.BytesIO(data))
    assert "Valuation" in reopened.sheetnames
