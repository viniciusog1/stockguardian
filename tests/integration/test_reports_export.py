"""Integração: export Excel dos relatórios."""

from __future__ import annotations

import io
import uuid

import pytest
from app.utils.excel import XLSX_MEDIA_TYPE
from httpx import AsyncClient
from openpyxl import load_workbook

pytestmark = pytest.mark.integration

PREFIX = "/api/v1"


async def _supplier(auth_client: AsyncClient) -> str:
    doc = str(uuid.uuid4().int)[:14]
    resp = await auth_client.post(f"{PREFIX}/suppliers", json={"name": "Forn", "document": doc})
    assert resp.status_code == 201, resp.text
    return resp.json()["id"]


async def _product(auth_client: AsyncClient, supplier_id: str, *, unit_price: str) -> str:
    sku = "X-" + uuid.uuid4().hex[:6]
    resp = await auth_client.post(
        f"{PREFIX}/products",
        json={"sku": sku, "name": "Prod", "supplier_id": supplier_id, "unit_price": unit_price},
    )
    assert resp.status_code == 201, resp.text
    return resp.json()["id"]


async def _move(auth_client: AsyncClient, pid: str, mtype: str, qty: int) -> None:
    resp = await auth_client.post(
        f"{PREFIX}/movements", json={"product_id": pid, "type": mtype, "quantity": qty}
    )
    assert resp.status_code == 201, resp.text


async def test_inventory_valuation_export(auth_client: AsyncClient) -> None:
    sid = await _supplier(auth_client)
    pid = await _product(auth_client, sid, unit_price="10.00")
    await _move(auth_client, pid, "in", 5)

    resp = await auth_client.get(f"{PREFIX}/reports/inventory-valuation/export")
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == XLSX_MEDIA_TYPE
    assert "attachment" in resp.headers["content-disposition"]
    assert ".xlsx" in resp.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb["Valuation"]
    values = [c.value for row in ws.iter_rows() for c in row]
    assert 50.0 in [v for v in values if isinstance(v, int | float)]


async def test_movements_summary_export(auth_client: AsyncClient) -> None:
    sid = await _supplier(auth_client)
    pid = await _product(auth_client, sid, unit_price="1.00")
    await _move(auth_client, pid, "in", 10)
    await _move(auth_client, pid, "out", 4)

    resp = await auth_client.get(
        f"{PREFIX}/reports/movements-summary/export", params={"product_id": pid}
    )
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == XLSX_MEDIA_TYPE

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb["Movimentações"]
    text = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    assert "in" in text and "out" in text and "adjustment" in text


async def test_export_requires_auth(client: AsyncClient) -> None:
    assert (await client.get(f"{PREFIX}/reports/inventory-valuation/export")).status_code == 401
    assert (await client.get(f"{PREFIX}/reports/movements-summary/export")).status_code == 401
