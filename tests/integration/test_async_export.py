"""Integração: export assíncrono — task direta + endpoints com fila fake."""

from __future__ import annotations

import io
import uuid
from typing import Any

import pytest
from app.dependencies.queue import get_report_queue
from app.schemas.report_job import ReportJobState
from app.utils.excel import ExportFile
from app.worker.tasks import generate_inventory_valuation_export
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncEngine, AsyncSession, async_sessionmaker

pytestmark = pytest.mark.integration

PREFIX = "/api/v1"


async def _supplier(auth_client: AsyncClient) -> str:
    doc = str(uuid.uuid4().int)[:14]
    resp = await auth_client.post(f"{PREFIX}/suppliers", json={"name": "Forn", "document": doc})
    assert resp.status_code == 201, resp.text
    return resp.json()["id"]


async def _product(auth_client: AsyncClient, supplier_id: str, *, unit_price: str) -> str:
    sku = "AS-" + uuid.uuid4().hex[:6]
    resp = await auth_client.post(
        f"{PREFIX}/products",
        json={"sku": sku, "name": "Prod", "supplier_id": supplier_id, "unit_price": unit_price},
    )
    assert resp.status_code == 201, resp.text
    return resp.json()["id"]


async def _move(auth_client: AsyncClient, pid: str, qty: int) -> None:
    resp = await auth_client.post(
        f"{PREFIX}/movements", json={"product_id": pid, "type": "in", "quantity": qty}
    )
    assert resp.status_code == 201, resp.text


async def test_task_generates_valid_xlsx(auth_client: AsyncClient, db_engine: AsyncEngine) -> None:
    sid = await _supplier(auth_client)
    pid = await _product(auth_client, sid, unit_price="10.00")
    await _move(auth_client, pid, 5)

    factory = async_sessionmaker(db_engine, class_=AsyncSession, expire_on_commit=False)
    result = await generate_inventory_valuation_export(
        {"session_factory": factory}, only_active=True
    )
    assert result["filename"].endswith(".xlsx")
    wb = load_workbook(io.BytesIO(result["content"]))
    values = [c.value for row in wb["Valuation"].iter_rows() for c in row]
    assert 50.0 in [v for v in values if isinstance(v, int | float)]


class _FakeQueue:
    """Fila em memória: simula enqueue/status/result sem ARQ/Redis."""

    def __init__(self) -> None:
        self.jobs: dict[str, dict[str, Any]] = {}

    async def enqueue_inventory_valuation(self, **_: object) -> str:
        jid = uuid.uuid4().hex
        self.jobs[jid] = {"state": ReportJobState.QUEUED, "export": None}
        return jid

    async def enqueue_movements_summary(self, **_: object) -> str:
        return await self.enqueue_inventory_valuation()

    async def get_status(self, job_id: str) -> ReportJobState:
        job = self.jobs.get(job_id)
        return job["state"] if job else ReportJobState.NOT_FOUND

    async def get_result(self, job_id: str) -> ExportFile | None:
        job = self.jobs.get(job_id)
        return job["export"] if job else None

    def complete(self, job_id: str) -> None:
        self.jobs[job_id]["state"] = ReportJobState.COMPLETE
        self.jobs[job_id]["export"] = ExportFile(
            filename="x.xlsx", media_type="application/octet-stream", content=b"PK-bytes"
        )


@pytest.fixture
def fake_queue(client: AsyncClient) -> _FakeQueue:
    fake = _FakeQueue()
    app = client._app  # type: ignore[attr-defined]
    app.dependency_overrides[get_report_queue] = lambda: fake
    return fake


async def test_enqueue_then_status_then_download(
    auth_client: AsyncClient, fake_queue: _FakeQueue
) -> None:
    accepted = await auth_client.post(f"{PREFIX}/reports/inventory-valuation/export-async")
    assert accepted.status_code == 202, accepted.text
    job_id = accepted.json()["job_id"]
    assert accepted.json()["status"] == "queued"

    # enquanto enfileirado: status queued, download 409
    st = await auth_client.get(f"{PREFIX}/reports/jobs/{job_id}")
    assert st.json()["status"] == "queued"
    assert (await auth_client.get(f"{PREFIX}/reports/jobs/{job_id}/download")).status_code == 409

    # conclui -> download entrega bytes
    fake_queue.complete(job_id)
    st2 = await auth_client.get(f"{PREFIX}/reports/jobs/{job_id}")
    assert st2.json()["status"] == "complete"
    dl = await auth_client.get(f"{PREFIX}/reports/jobs/{job_id}/download")
    assert dl.status_code == 200
    assert dl.content == b"PK-bytes"


async def test_unknown_job_is_404(auth_client: AsyncClient, fake_queue: _FakeQueue) -> None:
    assert (await auth_client.get(f"{PREFIX}/reports/jobs/nope")).status_code == 404
    assert (await auth_client.get(f"{PREFIX}/reports/jobs/nope/download")).status_code == 404


async def test_async_endpoints_require_auth(client: AsyncClient) -> None:
    assert (
        await client.post(f"{PREFIX}/reports/inventory-valuation/export-async")
    ).status_code == 401
    assert (await client.get(f"{PREFIX}/reports/jobs/x")).status_code == 401
