"""Serialização dos relatórios em planilhas Excel (.xlsx) com openpyxl.

Funções puras: recebem o schema do relatório e devolvem um Workbook. A rota
HTTP cuida do streaming/headers. Mantém a serialização testável sem DB.
"""

from __future__ import annotations

import io
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas.report import InventoryValuationReport, MovementsSummaryReport

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_MONEY_FORMAT = "#,##0.00"
_BOLD = Font(bold=True)


@dataclass(frozen=True)
class ExportFile:
    """Arquivo pronto para download: nome, media type e bytes."""

    filename: str
    media_type: str
    content: bytes


def _active_sheet(workbook: Workbook, title: str) -> Worksheet:
    ws = workbook.active
    if not isinstance(ws, Worksheet):  # pragma: no cover - Workbook() sempre cria a aba
        ws = workbook.create_sheet()
    ws.title = title
    return ws


def inventory_valuation_workbook(report: InventoryValuationReport) -> Workbook:
    wb = Workbook()
    ws = _active_sheet(wb, "Valuation")

    ws["A1"] = "Relatório de Valuation de Estoque"
    ws["A1"].font = _BOLD
    ws["A2"] = f"Gerado em: {report.generated_at.isoformat()}"

    header_row = 4
    headers = ["SKU", "Produto", "Quantidade", "Preço unitário", "Valor em estoque"]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = _BOLD

    row = header_row + 1
    for item in report.items:
        ws.cell(row=row, column=1, value=item.sku)
        ws.cell(row=row, column=2, value=item.name)
        ws.cell(row=row, column=3, value=item.quantity)
        price = ws.cell(row=row, column=4, value=item.unit_price)
        price.number_format = _MONEY_FORMAT
        value = ws.cell(row=row, column=5, value=item.stock_value)
        value.number_format = _MONEY_FORMAT
        row += 1

    row += 1  # linha em branco antes dos totais
    ws.cell(row=row, column=1, value="Total de produtos").font = _BOLD
    ws.cell(row=row, column=2, value=report.summary.total_products)
    row += 1
    ws.cell(row=row, column=1, value="Total de unidades").font = _BOLD
    ws.cell(row=row, column=2, value=report.summary.total_units)
    row += 1
    ws.cell(row=row, column=1, value="Valor total em estoque").font = _BOLD
    total = ws.cell(row=row, column=2, value=report.summary.total_value)
    total.number_format = _MONEY_FORMAT
    return wb


def movements_summary_workbook(report: MovementsSummaryReport) -> Workbook:
    wb = Workbook()
    ws = _active_sheet(wb, "Movimentações")

    ws["A1"] = "Relatório de Movimentações (resumo)"
    ws["A1"].font = _BOLD
    period_from = report.date_from.isoformat() if report.date_from else "início"
    period_to = report.date_to.isoformat() if report.date_to else "agora"
    ws["A2"] = f"Período: {period_from} → {period_to}"
    ws["A3"] = f"Gerado em: {report.generated_at.isoformat()}"

    header_row = 5
    headers = ["Tipo", "Qtd. de movimentações", "Quantidade total"]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = _BOLD

    row = header_row + 1
    for r in report.rows:
        ws.cell(row=row, column=1, value=r.type.value)
        ws.cell(row=row, column=2, value=r.movement_count)
        ws.cell(row=row, column=3, value=r.total_quantity)
        row += 1

    ws.cell(row=row, column=1, value="Total").font = _BOLD
    ws.cell(row=row, column=2, value=report.total_movements)
    return wb


def workbook_to_bytes(workbook: Workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def inventory_valuation_export_file(report: InventoryValuationReport) -> ExportFile:
    content = workbook_to_bytes(inventory_valuation_workbook(report))
    filename = f"inventory-valuation-{report.generated_at.date().isoformat()}.xlsx"
    return ExportFile(filename=filename, media_type=XLSX_MEDIA_TYPE, content=content)


def movements_summary_export_file(report: MovementsSummaryReport) -> ExportFile:
    content = workbook_to_bytes(movements_summary_workbook(report))
    filename = f"movements-summary-{report.generated_at.date().isoformat()}.xlsx"
    return ExportFile(filename=filename, media_type=XLSX_MEDIA_TYPE, content=content)
